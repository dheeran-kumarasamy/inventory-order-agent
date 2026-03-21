import io
import math
import re
from datetime import date
from typing import List, Optional, Set

import pandas as pd
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

HIGHLIGHT_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
ALT_ROW_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center")
MACH_HIGHLIGHT_COLS = set()
NAME_COL_W = 40    # Excel column width for product name columns
DATA_COL_W = 15    # Excel column width for stock/qty/sales columns
ROW_H_PER_LINE = 15  # Excel row height (pts) per wrapped line
NAME_COLS = {"Product Name", "RC Required", "RC Product Name", "Product name"}
DATA_COLS = {
    "Product Stock",
    "Avg Monthly Sales",
    "RC Stock",
    "Order",
    "Current Stock",
    "Stock",
    "M/C Order",
    "Manufacturing Order",
    "Avg Sales",
}

from sheets_loader import load_leadtime_from_sheet, load_sales_from_sheet, load_stock_from_sheet

NO_RC_PULLEY_TYPES = {
    "DOUBLE BASS FLAT PULLEY",
    "SINGLE BASS FLAT PULLEY",
    "HEAVY DOUBLE SIDE FLAT PULLEY",
    "HEAVY BASS FLAT PULLEY",
    "PADI PULLEY",
    "COUPLED FLANGE",
    "PLAIN PULLEY",
    "MUFF COUPLING",
    "STEERING WHEEL",
    "FAN / IMPELLER",
}

RC_MAPPING_RULES = [
    {
        "name": "DISC_TYPE",
        "match_all": ["DISC TYPE", "V-PULLEY"],
        "rc_tokens": ["DISC TYPE", "RC"],
        "expected_suffix": "DISC TYPE - RC",
    },
    {
        "name": "HALF_SOLID",
        "match_all": ["HALF SOLID", "V-PULLEY"],
        "rc_tokens": ["HALF SOLID", "RC"],
        "expected_suffix": "HALF SOLID - RC",
    },
    {
        "name": "HEAVY_BASS",
        "match_all": ["HEAVY BASS", "V-PULLEY"],
        "rc_tokens": ["HEAVY BASS", "RC"],
        "expected_suffix": "HEAVY BASS - RC",
    },
    {
        "name": "CENTRE_BASS",
        "match_any": ["CENTRE BASS", "CENTER BASS"],
        "match_all": ["V-PULLEY"],
        "rc_tokens": ["CB", "RC"],
        "expected_suffix": "CB - RC",
    },
    {
        "name": "LG",
        "match_all": ["LG V"],
        "rc_tokens": ["CB", "RC"],
        "expected_suffix": "CB - RC",
    },
    {
        "name": "LIGHT",
        "match_all": ["LIGHT V"],
        "rc_tokens": ["LIGHT", "RC"],
        "expected_suffix": "LIGHT - RC",
    },
    {
        "name": "SOLID",
        "match_all": ["SOLID", "V-PULLEY"],
        "rc_tokens": ["SOLID", "RC"],
        "expected_suffix": "SOLID - RC",
    },
    {
        "name": "HOLLOW",
        "match_all": ["HOLLOW", "V-PULLEY"],
        "rc_tokens": ["HOLLOW", "RC"],
        "expected_suffix": None,
    },
]


def calc_avg_sales(sales: pd.DataFrame) -> pd.DataFrame:
    grp = sales.groupby("ProductName")
    avg = (grp["Total"].sum() / grp["Month"].nunique()).reset_index()
    avg.columns = ["ProductName", "AvgMonthlySales"]
    avg["AvgDailySales"] = (avg["AvgMonthlySales"] / 30).round(3)
    return avg


def extract_parts(name: str):
    parts = re.split(r"\s*-\s*", name, maxsplit=1)
    prefix = parts[0].strip()
    tokens = [t.strip() for t in re.split(r"\s*[Xx]\s*", prefix)]
    od = tokens[0] if len(tokens) > 0 else ""
    grooves = tokens[1] if len(tokens) > 1 else ""
    return od, grooves


def extract_bass_size(product_name: str) -> Optional[str]:
    """Extract inch-size indicator from product name description (e.g. '5"' -> '5"')."""
    desc_parts = re.split(r"\s*-\s*", str(product_name), maxsplit=1)
    desc = desc_parts[1] if len(desc_parts) > 1 else str(product_name)
    match = re.search(r'(\d+)"', desc)
    if match:
        return f'{match.group(1)}"'
    return None


def normalize_product_name(name: str) -> str:
    normalized = str(name).upper().replace("×", "X")
    normalized = re.sub(r"\s*[Xx]\s*", " X ", normalized)
    normalized = re.sub(r"\s*-\s*", " - ", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    # Canonicalize "V PULLEY", "V-PULLEY", "V - PULLEY" → "V-PULLEY"
    normalized = re.sub(r"\bV[\s-]+PULLEY\b", "V-PULLEY", normalized)
    return normalized


def _contains_all(text: str, terms: List[str]) -> bool:
    return all(term in text for term in terms)


def _contains_any(text: str, terms: List[str]) -> bool:
    return any(term in text for term in terms)


def classify_pulley_type(product_name: str):
    normalized = normalize_product_name(product_name)

    for no_rc_name in NO_RC_PULLEY_TYPES:
        if no_rc_name in normalized:
            return "NO_RC"

    if "DISC TYPE" in normalized:
        return "DISC_TYPE"
    if "HALF SOLID" in normalized:
        return "HALF_SOLID"
    if "HEAVY BASS" in normalized:
        return "HEAVY_BASS"
    if "CENTRE BASS" in normalized or "CENTER BASS" in normalized:
        return "CENTRE_BASS"
    if "LIGHT V" in normalized:
        return "LIGHT"
    if "LG V" in normalized:
        return "LG"
    if "HOLLOW" in normalized:
        return "HOLLOW"
    if "SOLID" in normalized:
        return "SOLID"

    if "V-PULLEY" in normalized:
        return "UNKNOWN_V"

    return None


def get_mapping_rule(product_name: str):
    normalized = normalize_product_name(product_name)
    pulley_type = classify_pulley_type(product_name)
    if pulley_type in {"NO_RC", "UNKNOWN_V", None}:
        return pulley_type

    for rule in RC_MAPPING_RULES:
        if rule["name"] != pulley_type:
            continue
        if not _contains_all(normalized, rule.get("match_all", [])):
            continue
        match_any = rule.get("match_any")
        if match_any and not _contains_any(normalized, match_any):
            continue
        return rule

    return None


def build_expected_rc_name(od, grooves, rule):
    expected_suffix = rule.get("expected_suffix") if isinstance(rule, dict) else None
    if not (od and grooves and expected_suffix):
        return None
    return normalize_product_name(f"{od} X {grooves} - {expected_suffix}")


def find_rc(od, grooves, rule, rc_lookup):
    if not (od and grooves and isinstance(rule, dict)):
        return None

    candidate = build_expected_rc_name(od, grooves, rule)
    if candidate and candidate in rc_lookup:
        return rc_lookup[candidate]

    prefix = normalize_product_name(f"{od} X {grooves}")
    rc_tokens = rule.get("rc_tokens", [])
    matches = []
    for key, value in rc_lookup.items():
        normalized_key = normalize_product_name(key)
        if not normalized_key.startswith(prefix):
            continue
        if not all(token in normalized_key for token in rc_tokens):
            continue
        matches.append(value)

    if len(matches) == 1:
        return matches[0]
    return None


def resolve_pulley_rc(product_name, pulley_type, od, grooves, mapping_rule, rc_lookup):
    """Resolve RC with bass-size fallback rules for CENTRE_BASS/LG pulleys."""
    if not isinstance(mapping_rule, dict):
        return None, mapping_rule, "No mapping rule"

    if pulley_type not in ("CENTRE_BASS", "LG"):
        rc = find_rc(od, grooves, mapping_rule, rc_lookup)
        return rc, mapping_rule, "Standard rule"

    bass_size = extract_bass_size(product_name)
    if bass_size:
        sized_rule = {**mapping_rule, "rc_tokens": mapping_rule["rc_tokens"] + [f"{bass_size} BASS"]}
        rc = find_rc(od, grooves, sized_rule, rc_lookup)
        return rc, sized_rule, f"Used explicit bass size {bass_size}"

    # If bass size is not in product name: try base rule first, then 4" BASS fallback.
    rc = find_rc(od, grooves, mapping_rule, rc_lookup)
    if rc:
        return rc, mapping_rule, "No bass size in product name; matched RC without bass-size token"

    fallback_rule = {**mapping_rule, "rc_tokens": mapping_rule["rc_tokens"] + ['4" BASS']}
    rc = find_rc(od, grooves, fallback_rule, rc_lookup)
    if rc:
        return rc, fallback_rule, "No bass size in product name; matched 4\" BASS fallback"

    return None, fallback_rule, "No bass size in product name; tried no-bass and 4\" BASS fallback"


def _apply_excel_sheet(ws, df: pd.DataFrame, highlight_cols: Set[str]):
    cols = list(df.columns)
    col_idx = {c: i + 1 for i, c in enumerate(cols)}
    numeric_cols = {c for c in cols if pd.api.types.is_numeric_dtype(df[c])}

    for c in cols:
        letter = ws.cell(row=1, column=col_idx[c]).column_letter
        ws.column_dimensions[letter].width = NAME_COL_W if c in NAME_COLS else DATA_COL_W

    ws.append(cols)
    ws.row_dimensions[1].height = ROW_H_PER_LINE * 1.2

    for ri, (_, row) in enumerate(df.iterrows(), start=2):
        ws.append(list(row.values))
        max_lines = 1
        for c in cols:
            if c in NAME_COLS:
                val = str(row[c]) if pd.notna(row[c]) else ""
                lines = math.ceil(len(val) / NAME_COL_W) if val else 1
                max_lines = max(max_lines, lines)
        ws.row_dimensions[ri].height = max(ROW_H_PER_LINE, max_lines * ROW_H_PER_LINE)

    for rnum in range(1, len(df) + 2):
        is_header = rnum == 1
        is_even_data = not is_header and (rnum % 2 == 0)
        for c in cols:
            cidx = col_idx[c]
            cell = ws.cell(row=rnum, column=cidx)
            if is_header:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                if c in highlight_cols:
                    cell.fill = HIGHLIGHT_FILL
                elif is_even_data:
                    cell.fill = ALT_ROW_FILL
                else:
                    cell.fill = WHITE_FILL
                if c in numeric_cols:
                    cell.alignment = CENTER
                else:
                    cell.alignment = Alignment(horizontal="left", wrap_text=True, vertical="top")


def _build_excel(mach_out: pd.DataFrame, mfg_out: pd.DataFrame) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)

    ws_sum = wb.create_sheet("Summary")
    ws_sum.append(["Metric", "Value"])
    ws_sum.append(["Machining Items", len(mach_out)])
    ws_sum.append(["GP Items", len(mfg_out)])
    ws_sum.append(["Machining Units", int(mach_out["Order"].sum()) if len(mach_out) else 0])
    ws_sum.append(["GP Units", int(mfg_out["Order"].sum()) if len(mfg_out) else 0])

    _apply_excel_sheet(wb.create_sheet("Machining Orders"), mach_out, MACH_HIGHLIGHT_COLS)
    _apply_excel_sheet(wb.create_sheet("GP Orders"), mfg_out, set())

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_consolidated_excel(consolidated_out: pd.DataFrame) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)
    _apply_excel_sheet(wb.create_sheet("Consolidated order"), consolidated_out, set())

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _add_pdf_table(pdf: FPDF, title: str, df: pd.DataFrame, report_date: date, highlight_cols=None):
    highlight_cols = highlight_cols or set()
    pdf.add_page()
    if df.empty:
        pdf.set_font("Helvetica", "B", 18)
        pdf.cell(0, 12, "Inventory Order Report", new_x="LMARGIN", new_y="NEXT", align="C")
        pdf.set_font("Helvetica", "B", 13)
        pdf.cell(0, 10, title, new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 8, "No items.", new_x="LMARGIN", new_y="NEXT")
        return

    FONT_SIZE = 7
    LINE_H = 4
    PAD = 2
    cols = list(df.columns)
    avail = pdf.w - pdf.l_margin - pdf.r_margin
    numeric_cols = {c for c in cols if pd.api.types.is_numeric_dtype(df[c])}

    def _lines_needed(text, width, font_style=""):
        pdf.set_font("Helvetica", font_style, FONT_SIZE)
        if not text:
            return 1
        words = text.split()
        if not words:
            return 1
        lines, line = 1, ""
        for w in words:
            test = f"{line} {w}".strip()
            if pdf.get_string_width(test) > width - PAD:
                lines += 1
                line = w
            else:
                line = test
        return lines

    pdf.set_font("Helvetica", "", FONT_SIZE)
    col_widths = {}
    name_w = pdf.get_string_width("W" * 40) + PAD * 2
    data_w = pdf.get_string_width("0" * 15) + PAD * 2
    for c in cols:
        if c in DATA_COLS:
            col_widths[c] = data_w
        elif c in NAME_COLS:
            col_widths[c] = name_w

    fixed_total = sum(col_widths.get(c, 0) for c in cols)
    other_cols = [c for c in cols if c not in col_widths]
    remaining = avail - fixed_total
    if other_cols:
        per_other = max(remaining / len(other_cols), 25)
        for c in other_cols:
            col_widths[c] = per_other

    total_w = sum(col_widths.values())
    if total_w > avail and total_w > 0:
        scale = avail / total_w
        for c in cols:
            col_widths[c] = col_widths[c] * scale

    row_counter = [0]

    def _draw_page_header():
        pdf.set_font("Helvetica", "B", 18)
        pdf.cell(0, 12, "Inventory Order Report", new_x="LMARGIN", new_y="NEXT", align="C")
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 6, f"Generated: {report_date.strftime('%d %B %Y')}", new_x="LMARGIN", new_y="NEXT", align="C")
        pdf.ln(2)
        pdf.set_font("Helvetica", "B", 13)
        pdf.cell(0, 10, title, new_x="LMARGIN", new_y="NEXT")

    def _draw_row(values, is_header=False):
        x_start = pdf.l_margin
        y_start = pdf.get_y()
        style = "B" if is_header else ""
        pdf.set_font("Helvetica", style, FONT_SIZE)

        max_lines = 1
        for i, c in enumerate(cols):
            txt = str(values[i]) if values[i] is not None else ""
            w = col_widths[c]
            nl = _lines_needed(txt, w, style)
            if nl > max_lines:
                max_lines = nl
        row_h = max(max_lines * LINE_H, LINE_H)

        if y_start + row_h > pdf.h - 15:
            pdf.add_page()
            _draw_page_header()
            _draw_row(cols, is_header=True)
            y_start = pdf.get_y()

        if is_header:
            bg = (68, 114, 196)
        elif row_counter[0] % 2 == 0:
            bg = (217, 226, 243)
        else:
            bg = (255, 255, 255)

        if not is_header:
            row_counter[0] += 1

        for i, c in enumerate(cols):
            txt = str(values[i]) if values[i] is not None else ""
            w = col_widths[c]
            x = x_start + sum(col_widths[cols[j]] for j in range(i))
            is_highlighted = c in highlight_cols and not is_header

            if is_highlighted:
                pdf.set_fill_color(255, 230, 153)
            else:
                pdf.set_fill_color(*bg)
            pdf.rect(x, y_start, w, row_h, style="F")

            if is_header:
                pdf.set_xy(x, y_start)
                pdf.set_font("Helvetica", "B", FONT_SIZE)
                pdf.set_text_color(255, 255, 255)
                pdf.multi_cell(w, LINE_H, txt, border=0, align="C")
                pdf.set_text_color(0, 0, 0)
            elif c in DATA_COLS:
                y_off = (row_h - LINE_H) / 2
                pdf.set_xy(x, y_start + y_off)
                pdf.set_font("Helvetica", "", FONT_SIZE)
                pdf.cell(w, LINE_H, txt, border=0, align="C")
            else:
                pdf.set_xy(x, y_start)
                pdf.set_font("Helvetica", "", FONT_SIZE)
                pdf.multi_cell(w, LINE_H, txt, border=0, align="L")

            pdf.rect(x, y_start, w, row_h)

        pdf.set_y(y_start + row_h)

    _draw_page_header()
    _draw_row(cols, is_header=True)
    for _, row in df.iterrows():
        vals = [row[c] if pd.notna(row[c]) else "" for c in cols]
        _draw_row(vals)


def _build_pdf(mach_out: pd.DataFrame, mfg_out: pd.DataFrame, report_date: date) -> io.BytesIO:
    pdf = FPDF(orientation="L", format="A4")
    pdf.set_auto_page_break(auto=False)

    FONT_SIZE = 7
    LINE_H = 4
    PAD = 2

    # --- Summary Page ---
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 18)
    pdf.cell(0, 12, "Inventory Order Report", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 8, f"Generated: {report_date.strftime('%d %B %Y')}", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(8)

    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 10, "Summary", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    total_mach = int(mach_out["Order"].sum()) if len(mach_out) else 0
    total_mfg = int(mfg_out["Order"].sum()) if len(mfg_out) else 0
    for label, val in [
        ("Machining Items", len(mach_out)),
        ("GP Items", len(mfg_out)),
        ("Machining Units", total_mach),
        ("GP Units", total_mfg),
    ]:
        pdf.cell(80, 7, label, border=1)
        pdf.cell(40, 7, str(val), border=1, new_x="LMARGIN", new_y="NEXT")

    _add_pdf_table(pdf, "Machining Orders", mach_out, report_date, highlight_cols=MACH_HIGHLIGHT_COLS)
    _add_pdf_table(pdf, "GP Orders", mfg_out, report_date)

    buf = io.BytesIO()
    buf.write(pdf.output())
    buf.seek(0)
    return buf


def _build_consolidated_pdf(consolidated_out: pd.DataFrame, report_date: date) -> io.BytesIO:
    pdf = FPDF(orientation="L", format="A4")
    pdf.set_auto_page_break(auto=False)
    _add_pdf_table(pdf, "Consolidated order", consolidated_out, report_date)

    buf = io.BytesIO()
    buf.write(pdf.output())
    buf.seek(0)
    return buf


def generate_report(master_sheet_url, lt_url, mach_lead, mfg_lead):
    today = date.today()

    stock = load_stock_from_sheet(master_sheet_url)
    sales = load_sales_from_sheet(master_sheet_url)
    _ = load_leadtime_from_sheet(lt_url) if lt_url else None
    avg = calc_avg_sales(sales)

    below = stock[stock["StockLevel"] < stock["ReorderLevel"]].copy()
    rc_mask = below["ProductName"].str.contains(r"\bRC\b", case=False, na=False)
    mfg_df = below[rc_mask].copy()
    mach_df = below[~rc_mask].copy()

    all_rc = stock[stock["ProductName"].str.contains(r"\bRC\b", case=False, na=False)]
    rc_lookup = {normalize_product_name(n): n for n in all_rc["ProductName"]}
    rc_name_by_key = {n.upper(): n for n in all_rc["ProductName"]}
    rc_stock_lookup = {
        row["ProductName"].upper(): int(row["StockLevel"])
        for _, row in all_rc.iterrows()
    }
    remaining_rc_stock = dict(rc_stock_lookup)

    def enrich(df, lead):
        rows = []
        for _, row in df.iterrows():
            pname = row["ProductName"]
            stk = int(row["StockLevel"])
            reorder = int(row["ReorderLevel"])
            shortage = max(reorder - stk, 0)
            sm = avg[avg["ProductName"].str.upper() == pname.upper()]
            adaily = round(sm["AvgDailySales"].values[0], 3) if len(sm) else 0.0
            amonthly = round(sm["AvgMonthlySales"].values[0], 1) if len(sm) else 0.0
            suggested = max(int(adaily * lead) + shortage, 1)
            rows.append(
                {
                    "pname": pname,
                    "stk": stk,
                    "reorder": reorder,
                    "shortage": shortage,
                    "adaily": adaily,
                    "amonthly": amonthly,
                    "suggested": suggested,
                }
            )
        return rows

    mach_rows = []
    machining_shortfall_by_rc = {}
    for r in enrich(mach_df, mach_lead):
        pname = r["pname"]
        pulley_type = classify_pulley_type(pname)
        mapping_rule = get_mapping_rule(pname)
        is_mapped_pulley = isinstance(mapping_rule, dict)
        is_unknown_v_pulley = mapping_rule == "UNKNOWN_V"
        rc = None
        rc_available = None
        expected_rc = None
        od, grooves = "", ""
        effective_rule = mapping_rule
        resolution_note = ""
        if is_mapped_pulley:
            od, grooves = extract_parts(pname)
            rc, effective_rule, resolution_note = resolve_pulley_rc(
                pname,
                pulley_type,
                od,
                grooves,
                mapping_rule,
                rc_lookup,
            )
            expected_rc = build_expected_rc_name(od, grooves, effective_rule)
            if not rc and expected_rc:
                rc = expected_rc

        if rc:
            rc_key = rc.upper()
            rc_available = remaining_rc_stock.get(rc_key, 0)
            suggested = min(r["suggested"], rc_available)
            remaining_rc_stock[rc_key] = max(rc_available - suggested, 0)
            machining_shortfall = max(r["suggested"] - suggested, 0)
            if machining_shortfall:
                machining_shortfall_by_rc[rc_key] = machining_shortfall_by_rc.get(rc_key, 0) + machining_shortfall
            if rc_key not in rc_name_by_key:
                rc_name_by_key[rc_key] = rc
        else:
            suggested = 0 if (is_mapped_pulley or is_unknown_v_pulley) else r["suggested"]

        if is_mapped_pulley:
            rc_required = rc if rc else "No RC Found"
        elif mapping_rule == "NO_RC":
            rc_required = "N/A"
        elif is_unknown_v_pulley:
            rc_required = "No RC Mapping Rule"
        else:
            rc_required = "N/A"

        if mapping_rule == "NO_RC":
            rc_match_note = "No RC mapping required for this product type"
        elif is_unknown_v_pulley:
            rc_match_note = "Unknown V-pulley type; no mapping rule configured"
        elif is_mapped_pulley and rc:
            rc_match_note = (
                f"Matched with size prefix '{od} X {grooves}' and RC tokens "
                f"{', '.join(effective_rule.get('rc_tokens', []) if isinstance(effective_rule, dict) else [])}. "
                f"{resolution_note}"
            )
        elif is_mapped_pulley:
            rc_match_note = (
                f"No RC found for size prefix '{od} X {grooves}' with RC tokens "
                f"{', '.join(effective_rule.get('rc_tokens', []) if isinstance(effective_rule, dict) else [])}. "
                f"{resolution_note}"
            )
        else:
            rc_match_note = "Not a mapped pulley type"

        mach_rows.append(
            {
                "Product Name": pname,
                "Product Stock": r["stk"],
                "Avg Monthly Sales": r["amonthly"],
                "Detected Pulley Type": pulley_type if pulley_type else "N/A",
                "RC Required": rc_required,
                "RC Stock": rc_available if rc_available is not None else "N/A",
                "RC Match Note": rc_match_note,
                "Order": suggested,
            }
        )

    mfg_order_map = {}
    for r in enrich(mfg_df, mfg_lead):
        rc_key = r["pname"].upper()
        mfg_order_map[rc_key] = {
            "RC Product Name": r["pname"],
            "Current Stock": r["stk"],
            "Order": r["suggested"],
            "Avg Monthly Sales": r["amonthly"],
        }

    for rc_key, shortfall in machining_shortfall_by_rc.items():
        if rc_key in mfg_order_map:
            mfg_order_map[rc_key]["Order"] += shortfall
        else:
            mfg_order_map[rc_key] = {
                "RC Product Name": rc_name_by_key.get(rc_key, rc_key),
                "Current Stock": rc_stock_lookup.get(rc_key, 0),
                "Order": shortfall,
                "Avg Monthly Sales": 0.0,
            }

    mfg_rows = list(mfg_order_map.values())

    mach_out = pd.DataFrame(mach_rows).sort_values("Order", ascending=False)
    mfg_out = pd.DataFrame(mfg_rows)
    if not mfg_out.empty:
        mfg_out = mfg_out[mfg_out["Order"] > 0].sort_values("Order", ascending=False)
    consolidated_rows = []
    for _, row in mach_out.iterrows():
        consolidated_rows.append(
            {
                "Product name": row["Product Name"],
                "Stock": row["Product Stock"],
                "RC Stock": row["RC Stock"],
                "M/C Order": row["Order"],
                "Manufacturing Order": "",
                "Avg Sales": row["Avg Monthly Sales"],
            }
        )
    for _, row in mfg_out.iterrows():
        consolidated_rows.append(
            {
                "Product name": row["RC Product Name"],
                "Stock": row["Current Stock"],
                "RC Stock": "",
                "M/C Order": "",
                "Manufacturing Order": row["Order"],
                "Avg Sales": row["Avg Monthly Sales"] if "Avg Monthly Sales" in row else "",
            }
        )

    consolidated_out = pd.DataFrame(consolidated_rows)
    if not consolidated_out.empty:
        consolidated_out["_sort"] = pd.to_numeric(consolidated_out["M/C Order"], errors="coerce").fillna(0) + pd.to_numeric(consolidated_out["Manufacturing Order"], errors="coerce").fillna(0)
        consolidated_out = consolidated_out.sort_values(["_sort", "Product name"], ascending=[False, True]).drop(columns=["_sort"])

    excel_buf = _build_excel(mach_out, mfg_out)
    pdf_buf = _build_pdf(mach_out, mfg_out, today)
    consolidated_excel_buf = _build_consolidated_excel(consolidated_out)
    consolidated_pdf_buf = _build_consolidated_pdf(consolidated_out, today)
    return excel_buf, pdf_buf, consolidated_excel_buf, consolidated_pdf_buf, mach_out, mfg_out, consolidated_out, today


def generate_rc_mapping_report(master_sheet_url: str):
    """Generate a diagnostic mapping report for ALL stock products showing RC match status and reasoning."""
    stock = load_stock_from_sheet(master_sheet_url)

    all_rc = stock[stock["ProductName"].str.contains(r"\bRC\b", case=False, na=False)]
    rc_lookup = {normalize_product_name(n): n for n in all_rc["ProductName"]}
    rc_stock_lookup = {row["ProductName"].upper(): int(row["StockLevel"]) for _, row in all_rc.iterrows()}

    non_rc = stock[~stock["ProductName"].str.contains(r"\bRC\b", case=False, na=False)]

    rows = []
    for _, row in non_rc.iterrows():
        pname = row["ProductName"]
        stock_level = int(row["StockLevel"])
        pulley_type = classify_pulley_type(pname)
        mapping_rule = get_mapping_rule(pname)
        is_mapped_pulley = isinstance(mapping_rule, dict)

        rc_name = ""
        rc_stock_val = ""
        match_reason = ""

        if mapping_rule == "NO_RC":
            match_reason = f"No RC required — type '{pulley_type}' is in NO_RC_PULLEY_TYPES"
        elif mapping_rule == "UNKNOWN_V":
            match_reason = "V-pulley detected but no specific sub-type rule configured (e.g. SOLID, HOLLOW, CENTRE BASS, etc.)"
        elif mapping_rule is None:
            match_reason = "Not a V-pulley product — no RC mapping applies"
        elif is_mapped_pulley:
            od, grooves = extract_parts(pname)
            rc, effective_rule, resolution_note = resolve_pulley_rc(
                pname,
                pulley_type,
                od,
                grooves,
                mapping_rule,
                rc_lookup,
            )
            if rc:
                rc_name = rc
                rc_stock_val = rc_stock_lookup.get(rc.upper(), 0)
                match_reason = (
                    f"Matched via size prefix '{od} X {grooves}' + tokens: "
                    f"{', '.join(effective_rule.get('rc_tokens', []))}. {resolution_note}"
                )
            else:
                prefix = normalize_product_name(f"{od} X {grooves}")
                rc_tokens = effective_rule.get("rc_tokens", [])
                candidates = [
                    orig for key, orig in rc_lookup.items()
                    if key.startswith(prefix) and all(t in key for t in rc_tokens)
                ]
                if not candidates:
                    # Also check without bass size to diagnose ambiguity vs missing RC
                    base_tokens = mapping_rule.get("rc_tokens", [])
                    base_candidates = [
                        orig for key, orig in rc_lookup.items()
                        if key.startswith(prefix) and all(t in key for t in base_tokens)
                    ]
                    if base_candidates:
                        match_reason = (
                            f"No RC found for '{od} X {grooves}' + tokens {rc_tokens}. "
                            f"Found {len(base_candidates)} RC(s) without bass size filter: "
                            f"{'; '.join(str(c) for c in base_candidates[:3])}. {resolution_note}"
                        )
                    else:
                        match_reason = (
                            f"No RC in stock for size prefix '{od} X {grooves}' "
                            f"with tokens: {', '.join(rc_tokens)}. {resolution_note}"
                        )
                else:
                    match_reason = (
                        f"Ambiguous — {len(candidates)} RC products match "
                        f"'{od} X {grooves}' + tokens {rc_tokens}: "
                        f"{'; '.join(str(c) for c in candidates[:3])}. {resolution_note}"
                    )

        rows.append({
            "Product Name": pname,
            "Stock Level": stock_level,
            "Classified Type": pulley_type if pulley_type else "N/A",
            "Matched RC": rc_name if rc_name else "—",
            "RC Stock": rc_stock_val if rc_stock_val != "" else "—",
            "Mapping Reason": match_reason,
        })

    df = (
        pd.DataFrame(rows)
        .sort_values(["Classified Type", "Product Name"])
        .reset_index(drop=True)
    )

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("RC Mapping Audit")
    _apply_excel_sheet(ws, df, set())
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return df, buf
